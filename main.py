from openpyxl import load_workbook
from datetime import datetime


def main():
    Filename= input("Enter the name of Excel file: ")
    #Open an xlsx for reading
    wb = load_workbook(filename = Filename)
    #Get the current Active Sheet
    #You can also select a particular sheet
    #based on sheet name
    ws = wb["DayTrading Log"]
    #Open the csv file
    file = open("text.tlg", "r")
    lines = file.readlines()
    file.close()
    count = 0
    start_pos = int(input("Enter the starting position: ")) + 7
    for line in lines:
        if line != "" and line != " " and line != "\r" and line != "\n" and "|" in line:
            split = line.split("|")
            if split[0] == "STK_TRD":
                ws.cell(row=start_pos + count, column=3).value = split[2]
                ws.cell(row=start_pos + count, column=2).value = datetime.strptime(split[7], "%Y%m%d").date().strftime("%d/%m/%Y")
                ws.cell(row=start_pos + count, column=9).value = split[12]
                ws.cell(row=start_pos + count, column=10).value = split[10]
                count += 1
            print(line)
    #save the csb file
    wb.save(Filename)


if __name__ == '__main__':
    main()